# scripts/process_yilian.py
# -*- coding: utf-8 -*-

import os
from datetime import datetime, timedelta
import pandas as pd

# ---- 1) 工具函数 ----
def find_csv_files(repo_dir: str, skip_dir: str = "results"):
    csv_list = []
    for root, dirs, files in os.walk(repo_dir):
        if os.path.basename(root) == skip_dir:
            continue
        for f in files:
            if f.lower().endswith(".csv"):
                csv_list.append(os.path.join(root, f))
    return csv_list

def load_and_concat(csv_paths):
    frames = []
    for path in csv_paths:
        df = None
        for enc in ("gbk", "utf-8"):
            try:
                df = pd.read_csv(path, encoding=enc)
                break
            except Exception:
                continue
        if df is None:
            continue
        df.columns = [c.strip() for c in df.columns]
        # 兼容“基于哪条manifest分支\n”
        if {"项目名", "基于哪条manifest分支", "申请时间"}.issubset(df.columns):
            frames.append(df[["项目名", "基于哪条manifest分支", "申请时间"]].copy())
        elif {"项目名", "基于哪条manifest分支\n", "申请时间"}.issubset(df.columns):
            part = df[["项目名", "基于哪条manifest分支\n", "申请时间"]].rename(
                columns={"基于哪条manifest分支\n": "基于哪条manifest分支"}
            )
            frames.append(part.copy())
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

def prev_month_range(now_local: datetime):
    this_month_start = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_month_end = this_month_start - timedelta(seconds=1)
    prev_month_start = prev_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return prev_month_start, prev_month_end

# ---- 2) 主流程 ----
def main():
    # 仓库根路径和输出目录
    repo_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    results_dir = os.path.join(repo_dir, "results")
    os.makedirs(results_dir, exist_ok=True)

    # 汇集 CSV
    csv_files = find_csv_files(repo_dir, skip_dir="results")
    if not csv_files:
        print("No CSV files found.")
        # 仍然输出空文件，保持行为一致
        now_local = datetime.now()
        prev_start, _ = prev_month_range(now_local)
        month_str = prev_start.strftime("%Y-%m")
        empty = pd.DataFrame(columns=["项目名", "基于哪条manifest分支", "访问次数", "最近一次访问时间"])
        csv_path = os.path.join(results_dir, f"{month_str}_summary.csv")
        empty.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"Saved empty summary to {csv_path}")
        return

    df_all = load_and_concat(csv_files)
    if df_all.empty:
        print("No valid CSV data with required columns.")
        now_local = datetime.now()
        prev_start, _ = prev_month_range(now_local)
        month_str = prev_start.strftime("%Y-%m")
        empty = pd.DataFrame(columns=["项目名", "基于哪条manifest分支", "访问次数", "最近一次访问时间"])
        csv_path = os.path.join(results_dir, f"{month_str}_summary.csv")
        empty.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"Saved empty summary to {csv_path}")
        return

    # 解析时间（失败即 NaT）
    df_all["申请时间"] = pd.to_datetime(df_all["申请时间"], errors="coerce")

    # 计算上一自然月（按 runner 时间，若需洛杉矶时区可用 zoneinfo 做本地化）
    now_local = datetime.now()
    prev_start, prev_end = prev_month_range(now_local)
    month_str = prev_start.strftime("%Y-%m")

    # 过滤上一月数据
    m = (df_all["申请时间"] >= prev_start) & (df_all["申请时间"] <= prev_end)
    recent_df = df_all[m]

    # 如果上一月没有任何记录，输出空文件并结束
    if recent_df.empty:
        out = pd.DataFrame(columns=["项目名", "基于哪条manifest分支", "访问次数", "最近一次访问时间"])
        csv_path = os.path.join(results_dir, f"{month_str}_summary.csv")
        out.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"No data for previous month. Saved empty summary to {csv_path}")
        # 也生成空的 XLSX
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "Summary"
            ws.append(["项目名", "基于哪条manifest分支", "访问次数", "最近一次访问时间"])
            xlsx_path = os.path.join(results_dir, f"{month_str}_summary.xlsx")
            wb.save(xlsx_path)
            print(f"Saved empty Excel to {xlsx_path}")
        except Exception as e:
            print(f"Excel save skipped: {e}")
        return

    # 按项目+分支统计访问次数和最近时间
    summary = (
        recent_df.groupby(["项目名", "基于哪条manifest分支"])
        .agg(访问次数=("申请时间", "count"), 最近一次访问时间=("申请时间", "max"))
        .reset_index()
        .sort_values(["项目名", "访问次数"], ascending=[True, False])
    )

    # —— 关键修复：强制转为 datetime 再格式化，防止 .dt 报错 ——
    summary["最近一次访问时间"] = pd.to_datetime(summary["最近一次访问时间"], errors="coerce")
    summary["最近一次访问时间"] = summary["最近一次访问时间"].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")

    # 输出 CSV
    csv_path = os.path.join(results_dir, f"{month_str}_summary.csv")
    summary.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"Saved summary to {csv_path}")

    # 输出 Excel（自动列宽 + 冻结首行 + 简单样式）
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        for row in dataframe_to_rows(summary, index=False, header=True):
            ws.append(row)

        # 样式：表头加粗、底色，冻结首行
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 自动列宽
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

        # 添加表格样式
        tab = Table(displayName="SummaryTable", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        xlsx_path = os.path.join(results_dir, f"{month_str}_summary.xlsx")
        wb.save(xlsx_path)
        print(f"Saved Excel to {xlsx_path}")
    except Exception as e:
        print(f"Excel save skipped: {e}")

if __name__ == "__main__":
    main()
