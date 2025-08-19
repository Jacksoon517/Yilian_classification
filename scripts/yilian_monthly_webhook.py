# -*- coding: utf-8 -*-
"""
Yilian 月报：一体化脚本
- 扫描仓库 CSV（排除 results/），兼容两种列头：{"项目名","基于哪条manifest分支","申请时间"} 或其带换行版本
- 统计上一个自然月：访问次数、最近一次访问时间
- 生成 results/<YYYY-MM>_summary.csv / .xlsx
- 复制为 monthly/<YYYY-MM>.csv 并提交回仓库（需 GITHUB_TOKEN & contents:write 权限）
- 通过飞书“自定义机器人 Webhook”发送卡片（摘要 + 按钮：打开CSV / 打开多维表格）

环境变量（Actions 中设置）：
- FEISHU_WEBHOOK_URL       必填，自定义机器人 Webhook URL
- FEISHU_WEBHOOK_SECRET    选填，若机器人启用“签名校验”则必填
- FEISHU_BITABLE_URL       选填，多维表格的查看链接（按钮用）
- GITHUB_TOKEN             Actions 默认注入（需 workflow permissions: contents: write）
- GITHUB_REPOSITORY        默认注入，形如 org/repo
- GITHUB_REF_NAME          默认注入，当前分支；若无则脚本用 git 推断
"""

import os
import re
import json
import glob
import time
import hmac
import base64
import hashlib
import pathlib
import subprocess
from datetime import datetime, timedelta

import pandas as pd
import requests

# ----------------- 工具：时间与路径 ----------------- #
def prev_month_range(now_local: datetime):
    this_month_start = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_month_end = this_month_start - timedelta(seconds=1)
    prev_month_start = prev_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return prev_month_start, prev_month_end

def repo_root():
    return pathlib.Path(__file__).resolve().parents[1]

def ensure_dirs(*paths):
    for p in paths:
        pathlib.Path(p).mkdir(parents=True, exist_ok=True)

# ----------------- 第1部分：汇总上月 CSV ----------------- #
def find_csv_files(repo_dir: str, skip_dir: str = "results"):
    csv_list = []
    for root, dirs, files in os.walk(repo_dir):
        if pathlib.Path(root).name == skip_dir:
            continue
        for f in files:
            if f.lower().endswith(".csv"):
                csv_list.append(os.path.join(root, f))
    return csv_list

def load_and_concat(csv_paths):
    frames = []
    for path in csv_paths:
        df = None
        for enc in ("gbk", "utf-8", "utf-8-sig"):
            try:
                df = pd.read_csv(path, encoding=enc)
                break
            except Exception:
                continue
        if df is None:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        # 兼容“基于哪条manifest分支\n”
        if {"项目名", "基于哪条manifest分支", "申请时间"}.issubset(df.columns):
            frames.append(df[["项目名", "基于哪条manifest分支", "申请时间"]].copy())
        elif {"项目名", "基于哪条manifest分支\n", "申请时间"}.issubset(df.columns):
            part = df[["项目名", "基于哪条manifest分支\n", "申请时间"]].rename(
                columns={"基于哪条manifest分支\n": "基于哪条manifest分支"}
            )
            frames.append(part.copy())
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

def build_monthly_summary(df_all: pd.DataFrame, now_local: datetime):
    prev_start, prev_end = prev_month_range(now_local)
    month_str = prev_start.strftime("%Y-%m")
    if df_all.empty:
        out = pd.DataFrame(columns=["项目名", "基于哪条manifest分支", "访问次数", "最近一次访问时间"])
        return out, month_str

    df_all["申请时间"] = pd.to_datetime(df_all["申请时间"], errors="coerce")
    m = (df_all["申请时间"] >= prev_start) & (df_all["申请时间"] <= prev_end)
    recent_df = df_all[m]

    if recent_df.empty:
        out = pd.DataFrame(columns=["项目名", "基于哪条manifest分支", "访问次数", "最近一次访问时间"])
        return out, month_str

    summary = (
        recent_df.groupby(["项目名", "基于哪条manifest分支"])
        .agg(访问次数=("申请时间", "count"), 最近一次访问时间=("申请时间", "max"))
        .reset_index()
        .sort_values(["项目名", "访问次数"], ascending=[True, False])
    )
    summary["最近一次访问时间"] = pd.to_datetime(summary["最近一次访问时间"], errors="coerce")
    summary["最近一次访问时间"] = summary["最近一次访问时间"].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
    return summary, month_str

def save_outputs(summary: pd.DataFrame, month: str, out_dir: pathlib.Path):
    ensure_dirs(out_dir)
    csv_path = out_dir / f"{month}_summary.csv"
    summary.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # 尝试额外导出 Excel（可选）
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook(); ws = wb.active; ws.title = "Summary"
        for row in dataframe_to_rows(summary, index=False, header=True):
            ws.append(row)

        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        for cell in ws[1]:
            cell.font = Font(bold=True); cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

        tab = Table(displayName="SummaryTable", ref=ws.dimensions)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        xlsx_path = out_dir / f"{month}_summary.xlsx"
        ws.add_table(tab); wb.save(xlsx_path)
    except Exception as e:
        print(f"[warn] Excel save skipped: {e}")

    return str(csv_path)

# ----------------- 第2部分：提交 monthly/<YYYY-MM>.csv ----------------- #
def git_commit_and_push_monthly(csv_src: str, month: str):
    repo = repo_root()
    monthly_dir = repo / "monthly"
    ensure_dirs(monthly_dir)
    dst = monthly_dir / f"{month}.csv"
    pathlib.Path(dst).write_bytes(pathlib.Path(csv_src).read_bytes())

    def sh(cmd):
        return subprocess.check_output(cmd, cwd=repo, text=True)

    # 配置 git 用户
    subprocess.run(["git", "config", "user.name",  "github-actions[bot]"], cwd=repo, check=False)
    subprocess.run(["git", "config", "user.email", "41898282+github-actions[bot]@users.noreply.github.com"], cwd=repo, check=False)

    # 提交变更
    subprocess.run(["git", "add", str(dst.relative_to(repo))], cwd=repo, check=True)
    subprocess.run(["git", "commit", "-m", f"monthly: add {month}.csv [skip ci]"], cwd=repo, check=False)
    subprocess.run(["git", "push"], cwd=repo, check=False)

    # 生成原始下载链接
    branch = os.environ.get("GITHUB_REF_NAME")
    if not branch:
        try:
            branch = sh(["git", "rev-parse", "--abbrev-ref", "HEAD"]).strip()
        except Exception:
            branch = "main"
    repo_slug = os.environ.get("GITHUB_REPOSITORY", "")
    csv_url = f"https://raw.githubusercontent.com/{repo_slug}/{branch}/monthly/{month}.csv"
    return csv_url

# ----------------- 第3部分：Webhook 卡片 ----------------- #
def sign_headers(secret: str):
    ts = str(int(time.time()))
    string_to_sign = f"{ts}\n{secret}"
    digest = hmac.new(secret.encode("utf-8"), string_to_sign.encode("utf-8"), hashlib.sha256).digest()
    sign = base64.b64encode(digest).decode("utf-8")
    return ts, sign

def build_card_from_summary(df: pd.DataFrame, month: str, csv_url: str, bitable_url: str = ""):
    total = len(df)
    # Top 5：若有“访问次数”按其排序，否则取前5行
    top = df.copy()
    if "访问次数" in top.columns:
        try:
            top["访问次数"] = pd.to_numeric(top["访问次数"], errors="coerce").fillna(0).astype(int)
        except Exception:
            pass
        top = top.sort_values("访问次数", ascending=False)
    top_rows = top.head(5).to_dict("records")

    # 组装 Markdown 列表
    md_lines = []
    for i, r in enumerate(top_rows, 1):
        if "项目名" in r:
            line = f"{i}. {r.get('项目名', '')}"
            if "基于哪条manifest分支" in r and r.get("基于哪条manifest分支", ""):
                line += f"（{r['基于哪条manifest分支']}）"
            if "访问次数" in r and r.get("访问次数", "") != "":
                line += f" · {r['访问次数']} 次"
            md_lines.append(line)
        elif "标题" in r:
            # 兼容另一类数据：标题/链接/分类/发布时间
            title = str(r.get("标题", ""))[:120]
            link  = str(r.get("链接", ""))
            md_lines.append(f"{i}. [{title}]({link})" if link.startswith("http") else f"{i}. {title}")
    md = "\n".join(md_lines) if md_lines else "（无数据）"

    actions = []
    if csv_url:
        actions.append({
            "tag": "button",
            "text": {"tag": "plain_text", "content": "打开CSV"},
            "type": "primary",
            "url": csv_url
        })
    if bitable_url:
        actions.append({
            "tag": "button",
            "text": {"tag": "plain_text", "content": "打开多维表格"},
            "url": bitable_url
        })

    card = {
      "config": {"wide_screen_mode": True, "enable_forward": True},
      "header": {"template": "blue", "title": {"tag": "plain_text", "content": f"Yilian 月报 · {month}"}},
      "elements": [
        {"tag": "div",
         "fields": [
            {"is_short": True, "text": {"tag": "lark_md", "content": f"**本月条数**\n{total}"}},
            {"is_short": True, "text": {"tag": "lark_md", "content": f"**生成时间**\n{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}" }}
         ]},
        {"tag": "div", "text": {"tag": "lark_md", "content": f"**Top 5**\n{md}"}},
      ] + ([{"tag": "action", "actions": actions}] if actions else [])
    }
    return card

def send_webhook_card(card: dict, webhook: str, secret: str = ""):
    body = {"msg_type": "interactive", "card": card}
    if secret:
        ts, sig = sign_headers(secret)
        body.update({"timestamp": ts, "sign": sig})
    r = requests.post(webhook, json=body, timeout=30)
    r.raise_for_status()
    return r.text

# ----------------- 主流程 ----------------- #
def main():
    WEBHOOK = os.environ.get("FEISHU_WEBHOOK_URL")
    assert WEBHOOK, "缺少 FEISHU_WEBHOOK_URL"

    SECRET  = os.environ.get("FEISHU_WEBHOOK_SECRET", "")
    BITABLE_URL = os.environ.get("FEISHU_BITABLE_URL", "")

    root = repo_root()
    results_dir = root / "results"
    ensure_dirs(results_dir)

    # 1) 汇总
    csv_files = find_csv_files(str(root), skip_dir="results")
    df_all = load_and_concat(csv_files)
    summary, month = build_monthly_summary(df_all, datetime.now())
    csv_path = save_outputs(summary, month, results_dir)

    # 2) 提交 monthly/ 并生成原始链接
    csv_url = git_commit_and_push_monthly(csv_path, month)

    # 3) 通过 Webhook 发卡片
    card = build_card_from_summary(summary, month, csv_url, BITABLE_URL)
    resp = send_webhook_card(card, WEBHOOK, SECRET)
    print("Webhook OK:", resp)

if __name__ == "__main__":
    main()
